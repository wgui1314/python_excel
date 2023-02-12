
from fastapi import FastAPI
from openpyxl import load_workbook
from api_model import labour
import load_config.load_config as LoadConfing
app = FastAPI()
a = load_config.LoadConfig()
import uvicorn
class GenerateTemplete:

    def __init__(self,params):
      # self.product_list = params["product_list"]
      # self.labour_categories = params["product_list"]
      self.open_worksheet()

    def operate_worksheet(self):
        pass
    


    def open_worksheet(self):
        templete_workbook = load_workbook(filename="./config/labour_templete.xlsx")

        self.write_product(templete_workbook,[12313,2131])
        self.write_firm_info(templete_workbook, {})
        self.wirte_labour_category(templete_workbook,[12313,2131])
        self.wirte_product_category(templete_workbook,[12313,2131])
        templete_workbook.save(filename="./config/wb2.xlsx")

    def write_product(self,workbook,product_list):
         curret_sheet = workbook.worksheets[0]
         for index,product in enumerate(product_list):
            curret_sheet.cell(row=index+2, column=1, value="12312") #产品ID
            curret_sheet.cell(row=index+2, column=2, value="纳税人识别号")  # 产品名称
            curret_sheet.cell(row=index+2, column=3, value="金额")  # 劳务分类
            curret_sheet.cell(row=index + 2, column=4, value="金额")  # 产品分类
            curret_sheet.cell(row=index + 2, column=5, value="金额")  # 价格构成
            curret_sheet.cell(row=index + 2, column=6, value="金额")  # 劳务价格
            curret_sheet.cell(row=index + 2, column= 7, value="金额")  # 单位
            curret_sheet.cell(row=index + 2, column= 8, value="金额")  # 税率
            curret_sheet.cell(row=index + 2, column=9, value="金额")  # 工作内容

    def write_firm_info(self,workbook,firm_info):
         curret_sheet = workbook.worksheets[1]
         curret_sheet.cell(row=2, column=1, value="公司名称") 
         curret_sheet.cell(row=2, column=2, value="加密密钥") 

    def wirte_labour_category(self,workbook,labour_categories):
         curret_sheet = workbook.worksheets[2]
         for index,labour_category in enumerate(labour_categories): 
            curret_sheet.cell(row=index+1, column=1, value="劳务分类")
         curret_sheet.sheet_state = 'hidden'
    def wirte_product_category(self,workbook,product_categories):
         curret_sheet = workbook.worksheets[3]
         for index,product_category in enumerate(product_categories): 
            curret_sheet.cell(row=index+1, column=1, value="产品分类")
         curret_sheet.sheet_state = 'hidden'



@app.post("/export_labor_excel")
def read_root(params:labour.DefaultTemplete):
    GenerateTemplete(params)
    return {"Hello": "World"}

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)